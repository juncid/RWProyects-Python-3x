#pip install openpyxl
import openpyxl as oxl

excel = oxl.load_workbook('emails.xlsx')
data = excel['Sheet1']

last_row = data.max_row

emails = []
names = []

for i in range(2, last_row+1):
    flag = data.cell(row=i, column=3).value
    if flag !='N':
        names.append(data.cell(row=i, column=1).value)

for i in range(2, last_row+1):
    flag = data.cell(row=i, column=3).value
    if flag !='N':
        emails.append(data.cell(row=i, column=2).value)

print(names)
print(emails)